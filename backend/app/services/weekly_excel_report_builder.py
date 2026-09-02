import io
import os
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill  # type: ignore
from openpyxl.utils.dataframe import dataframe_to_rows  # type: ignore
from openpyxl.drawing.image import Image # type: ignore
from typing import Any

def _write_table_to_sheet(worksheet, start_row: int, title: str, columns: list[str], data: list[dict]) -> int:
    # Title
    slate_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    thick_side = Side(style='thick')
    thin_side = Side(style='thin')
    
    worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(columns))
    title_cell = worksheet.cell(row=start_row, column=1)
    title_cell.value = title
    title_cell.font = Font(name="Calibri", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = slate_fill
    worksheet.row_dimensions[start_row].height = 30
    
    for c in range(1, len(columns) + 1):
        cell = worksheet.cell(row=start_row, column=c)
        cell.border = Border(top=thick_side, bottom=thick_side, left=thin_side, right=thin_side)
    
    # Headers
    header_row_1 = start_row + 1
    header_row_2 = start_row + 2
    worksheet.row_dimensions[header_row_1].height = 82.5
    for col_idx, col_name in enumerate(columns, start=1):
        cell = worksheet.cell(row=header_row_1, column=col_idx)
        if col_idx < 18:
            worksheet.merge_cells(start_row=header_row_1, start_column=col_idx, end_row=header_row_2, end_column=col_idx)
            cell.value = col_name
            cell.alignment = Alignment(textRotation=90, horizontal="center", vertical="center", wrap_text=True)
        elif col_idx == 18:
            worksheet.merge_cells(start_row=header_row_1, start_column=18, end_row=header_row_1, end_column=20)
            cell.value = "Exemption Permits"
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        cell.font = Font(name="Calibri", bold=True, size=14)

    sub_headers = ["Not Weighd (E)", "Weighed(F)", "Total"]
    for i, sub_header in enumerate(sub_headers):
        cell = worksheet.cell(row=header_row_2, column=18 + i)
        cell.value = sub_header
        cell.font = Font(name="Calibri", bold=True, size=14)
        cell.alignment = Alignment(textRotation=90, horizontal="center", vertical="center", wrap_text=True)

    df = pd.DataFrame(data, columns=columns)
    
    totals: dict[str, Any] = {"DATE": "Total"}
    for col in columns[1:]:
        totals[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).sum()
        totals[col] = int(totals[col]) if pd.notnull(totals[col]) else ""
        
    df = pd.concat([df, pd.DataFrame([totals])], ignore_index=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    slate_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")

    current_row = header_row_2 + 1
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=current_row):
        is_total_row = (r_idx == current_row + len(df) - 1)
        for c_idx, value in enumerate(row, start=1):
            cell = worksheet.cell(row=r_idx, column=c_idx)
            
            if pd.isna(value):
                cell_value = ""
            else:
                cell_value = value
                
            # Date formatting
            if c_idx == 1 and cell_value != "Total" and cell_value:
                try:
                    cell_value = datetime.strptime(str(cell_value), "%Y-%m-%d").strftime("%d-%b-%Y")
                except ValueError:
                    pass
                    
            cell.value = cell_value
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if is_total_row:
                cell.font = Font(name="Calibri", bold=True, size=14)
                cell.fill = slate_fill
                cell.border = Border(top=thick_side, bottom=thick_side, left=thin_side, right=thin_side)
            else:
                cell.border = thin_border
                if c_idx == 1:
                    cell.font = Font(name="Calibri", bold=True, size=14)
                    cell.fill = slate_fill
                else:
                    cell.font = Font(name="Calibri", size=14)
                    # Apply yellow only to H, C, Q, T (which correspond to c_idx 2, 3, 4, 9)
                    if c_idx in [2, 3, 4, 9]:
                        cell.fill = yellow_fill
                
    for r in range(header_row_1, header_row_2 + 1):
        for c in range(1, 21):
            cell = worksheet.cell(row=r, column=c)
            top_style = thick_side if r == header_row_1 else thin_side
            bottom_style = thick_side if r == header_row_2 else thin_side
            
            # Preserve existing properties
            current_left = cell.border.left if cell.border else thin_side
            current_right = cell.border.right if cell.border else thin_side
            
            cell.border = Border(top=top_style, bottom=bottom_style, left=current_left, right=current_right)
            
    return current_row + len(df)


def build_weekly_excel_report(
    weekly_data_by_bound: dict[str, list[dict]],
    weekly_data_combined: list[dict],
    start_date: str,
    end_date: str,
    station: str,
    prepared_by: str,
    approved_by: str,
) -> io.BytesIO:
    columns = [
        "DATE", "HSWIM Total (H)", "Called in (C)", "Cleared by HSWIM\n(Q) = (H-C)",
        "Weighed Scale (N)= D+S", "Manually Weighed (M)", "Total Weighed (X) = (N+M)",
        "Total Traffic Census =(K)", "Total Traffic (T) = (Q+X+K+E)", "Total Overloaded (Y)=(A+G+P)",
        "Impounded & Prohibited (P) = (Z+R)", "Warned (A)", "Prohibited & Charged (Z)",
        "Special Released (G)", "Redistributed (R)", "Cases Cleared in Court (B)",
        "Transgressions (L)", "Not Weighd (E)", "Weighed(F)", "Total"
    ]

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Create a dummy DataFrame to initialize sheet
        pd.DataFrame().to_excel(writer, index=False, startrow=0, header=False, sheet_name="JUJA WB WEEKLY REPORT ")
        worksheet = writer.sheets["JUJA WB WEEKLY REPORT "]
        
        # Add Logo
        logo_path = os.path.join(os.path.dirname(__file__), "../assets/logo.png")
        if os.path.exists(logo_path):
            worksheet.merge_cells("A1:C2")
            img = Image(logo_path)
            img.width = 120
            img.height = 40
            worksheet.add_image(img, "A1")

        current_row = 4
        for bound_name, data in weekly_data_by_bound.items():
            title = f"{station.upper()} WEIGHBRIDGE {bound_name.upper()} WEEKLY SUMMARY REPORT"
            current_row = _write_table_to_sheet(worksheet, current_row, title, columns, data)

        title_combined = f"{station.upper()} WEIGHBRIDGE WEEKLY TOTAL SUMMARY REPORT"
        current_row = _write_table_to_sheet(worksheet, current_row, title_combined, columns, weekly_data_combined)

        # Column widths
        worksheet.column_dimensions["A"].width = 15
        for col_letter in [chr(i) for i in range(ord('B'), ord('U'))]:
            worksheet.column_dimensions[col_letter].width = 12

        # Signatures
        prep_cell = worksheet.cell(row=current_row + 2, column=1)
        prep_cell.value = f"PREPARED BY: {prepared_by.upper()}"
        prep_cell.font = Font(name="Calibri", bold=True, size=14)
        
        app_cell = worksheet.cell(row=current_row + 4, column=1)
        app_cell.value = f"APPROVED BY: {approved_by.upper()}"
        app_cell.font = Font(name="Calibri", bold=True, size=14)
        
        # Footer Table
        footer_row = current_row + 6
        slate_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        worksheet.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=8)
        worksheet.merge_cells(start_row=footer_row, start_column=9, end_row=footer_row, end_column=14)
        worksheet.merge_cells(start_row=footer_row, start_column=15, end_row=footer_row, end_column=20)
        
        f1 = worksheet.cell(row=footer_row, column=1)
        f1.value = f"{station.upper()} WEIGHBRIDGE WEEKLY SUMMARY REPORT"
        
        f2 = worksheet.cell(row=footer_row, column=9)
        f2.value = "KeNHA/WB/MTCE/4339/2025"
        
        f3 = worksheet.cell(row=footer_row, column=15)
        f3.value = f"FOR DATES: {start_date} TO {end_date}"
        
        for c in range(1, 21):
            cell = worksheet.cell(row=footer_row, column=c)
            cell.fill = slate_fill
            cell.border = Border(
                top=Side(style='thin'),
                bottom=Side(style='thin'),
                left=Side(style='thin') if c == 1 else None,
                right=Side(style='thin') if c == 20 else None
            )
            if c in [1, 9, 15]:
                cell.font = Font(name="Calibri", bold=True, size=14)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    buffer.seek(0)
    return buffer
