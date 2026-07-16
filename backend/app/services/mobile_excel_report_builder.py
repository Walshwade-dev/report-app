import io
from copy import copy
from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import CharacterProperties, Paragraph, ParagraphProperties, RichTextProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from app.services.daily_hour_processor import HOURS
from app.services.mobile_report_processor import summarize_mobile_report


FONT_NAME = "Arial"
DETAIL_FONT_NAME = "Calibri"
DATE_FORMAT = "dd-mmm-yy"

THIN_SIDE = Side(style="thin", color="000000")
MEDIUM_SIDE = Side(style="medium", color="000000")
THIN_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)
MEDIUM_BORDER = Border(
    left=MEDIUM_SIDE,
    right=MEDIUM_SIDE,
    top=MEDIUM_SIDE,
    bottom=MEDIUM_SIDE,
)
THICK_SIDE = Side(style="thick", color="000000")
THICK_BORDER = Border(
    left=THICK_SIDE,
    right=THICK_SIDE,
    top=THICK_SIDE,
    bottom=THICK_SIDE,
)
MEDIUM_BORDER_THIN_BOTTOM = Border(
    left=MEDIUM_SIDE,
    right=MEDIUM_SIDE,
    top=MEDIUM_SIDE,
    bottom=THIN_SIDE,
)

NO_FILL = PatternFill(fill_type=None)
RED_FILL = PatternFill("solid", fgColor="FFFF0000")
LIGHT_GREY_FILL = PatternFill("solid", fgColor="F2F2F2")
DARK_BLUE_LINE = "1F4E79"
MAROON_LINE = "800000"


SUMMARY_COLUMN_WIDTHS = {
    "B": 5.43,
    "C": 12.29,
    "D": 12.0,
    "E": 10.86,
    "F": 10.14,
    "G": 11.29,
    "H": 12.0,
    "I": 11.86,
    "J": 11.43,
    "K": 12.86,
    "L": 13.29,
    "M": 12.43,
    "N": 13.14,
    "O": 8.43,
    "P": 8.43,
    "Q": 8.43,
    "R": 8.43,
    "S": 8.43,
    "T": 8.43,
    "U": 8.43,
    "V": 8.43,
    "W": 8.43,
    "X": 8.43,
}

DETAIL_COLUMN_WIDTHS = {
    "A": 2.0,
    "B": 11.43,
    "C": 16.43,
    "D": 43.29,
    "E": 12.14,
    "F": 27.29,
    "G": 19.43,
    "H": 21.71,
    "I": 13.57,
    "J": 31.29,
    "K": 65.71,
    "L": 42.71,
    "M": 23.14,
    "N": 132.86,
}


def _font(
    size: float = 11,
    bold: bool = False,
    *,
    name: str = FONT_NAME,
) -> Font:
    return Font(name=name, size=size, bold=bold, color="000000")


def _alignment(
    horizontal: str | None = "center",
    vertical: str | None = "center",
    wrap_text: bool | None = True,
) -> Alignment:
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)


def _set_cell(
    ws: Worksheet,
    coordinate: str,
    value: Any = None,
    *,
    size: float = 11,
    bold: bool = False,
    font_name: str = FONT_NAME,
    fill: PatternFill | None = NO_FILL,
    border: Border | None = THIN_BORDER,
    horizontal: str | None = "center",
    vertical: str | None = "center",
    wrap_text: bool | None = True,
    number_format: str = "General",
) -> None:
    cell: Any = ws[coordinate]
    cell.value = value
    cell.font = _font(size=size, bold=bold, name=font_name)
    cell.alignment = _alignment(horizontal, vertical, wrap_text)
    if fill is not None:
        cell.fill = copy(fill)
    if border is not None:
        cell.border = copy(border)
    cell.number_format = number_format


def _style_range(
    ws: Worksheet,
    cell_range: str,
    *,
    size: float = 11,
    bold: bool = False,
    font_name: str = FONT_NAME,
    fill: PatternFill | None = NO_FILL,
    border: Border | None = THIN_BORDER,
    horizontal: str | None = "center",
    vertical: str | None = "center",
    wrap_text: bool | None = True,
    number_format: str = "General",
) -> None:
    for row in ws[cell_range]:
        for cell in row:
            cell.font = _font(size=size, bold=bold, name=font_name)
            cell.alignment = _alignment(horizontal, vertical, wrap_text)
            if fill is not None:
                cell.fill = copy(fill)
            if border is not None:
                cell.border = copy(border)
            cell.number_format = number_format


def _merge_and_set(
    ws: Worksheet,
    cell_range: str,
    value: Any = None,
    **style_kwargs,
) -> None:
    _style_range(ws, cell_range, **style_kwargs)
    ws.merge_cells(cell_range)
    _set_cell(ws, cell_range.split(":", 1)[0], value, **style_kwargs)


def _set_detail_text_cell(
    ws: Worksheet,
    coordinate: str,
    value: Any = None,
    *,
    border: Border | None = THIN_BORDER,
    vertical: str | None = "center",
    number_format: str = "General",
) -> None:
    _set_cell(
        ws,
        coordinate,
        value,
        font_name=DETAIL_FONT_NAME,
        border=border,
        horizontal="left",
        vertical=vertical,
        wrap_text=False,
        number_format=number_format,
    )


def _set_detail_center_cell(
    ws: Worksheet,
    coordinate: str,
    value: Any = None,
    *,
    border: Border | None = THIN_BORDER,
    number_format: str = "General",
    size: float = 11,
) -> None:
    _set_cell(
        ws,
        coordinate,
        value,
        size=size,
        font_name=DETAIL_FONT_NAME,
        border=border,
        horizontal="center",
        vertical="center",
        wrap_text=False,
        number_format=number_format,
    )


def _chart_axis_text_properties() -> RichText:
    return RichText(
        bodyPr=RichTextProperties(
            rot=-60000000,
            spcFirstLastPara=True,
            vertOverflow="ellipsis",
            vert="horz",
            wrap="square",
            anchor="ctr",
            anchorCtr=True,
        ),
        p=[
            Paragraph(
                pPr=ParagraphProperties(
                    defRPr=CharacterProperties(
                        sz=900,
                        b=False,
                        i=False,
                        u="none",
                        strike="noStrike",
                        kern=1200,
                        baseline=0,
                    )
                )
            )
        ],
    )


def _note_key_value(key: str, value: str) -> CellRichText:
    return CellRichText(
        TextBlock(InlineFont(b=True, color="000000"), key),
        TextBlock(InlineFont(b=False, color="000000"), f" = {value}"),
    )


def _merge_and_set_note_row(
    ws: Worksheet,
    cell_range: str,
    key: str,
    value: str,
    *,
    fill: PatternFill | None = NO_FILL,
) -> None:
    _style_range(
        ws,
        cell_range,
        fill=fill,
        border=MEDIUM_BORDER,
        horizontal="left",
        wrap_text=False,
    )
    ws.merge_cells(cell_range)
    cell: Any = ws[cell_range.split(":", 1)[0]]
    cell.value = _note_key_value(key, value)


def _plain_int(value: Any) -> int | None:
    number = pd.to_numeric(str(value).replace(",", ""), errors="coerce")
    if pd.isna(number):
        return None
    return int(number)


def _upper(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, (list, tuple)):
        return " / ".join(_upper(item) for item in value if _upper(item))

    text = str(value).strip()
    if not text:
        return ""
    return text.upper()


def _manual_source(session) -> dict[str, Any]:
    mobile = session.manual_inputs.get("mobile_report")
    if isinstance(mobile, dict):
        return mobile

    extra = session.manual_inputs.get("extra")
    if isinstance(extra, dict) and isinstance(extra.get("mobile_report"), dict):
        return extra["mobile_report"]

    return {}


def _manual_value(session, *keys: str, default: Any = "") -> Any:
    mobile = _manual_source(session)
    for source in (mobile, session.manual_inputs):
        for key in keys:
            value = source.get(key)
            if value not in (None, ""):
                return value
    return default


def _manual_shifts(session) -> list[dict[str, Any]]:
    mobile = _manual_source(session)
    shifts = mobile.get("shifts")
    if isinstance(shifts, list):
        valid_shifts = [shift for shift in shifts if isinstance(shift, dict)]
        if valid_shifts:
            return valid_shifts

    return [
        {
            "label": "Shift 1",
            "start_time": "0000",
            "end_time": "0000",
            "danka_staff": _manual_value(
                session,
                "danka_staff",
                "danka_officers",
                "computer_operator",
                "computer_operators",
            ),
            "police_officers": _manual_value(session, "police_officers", "police"),
            "mobile_vehicle": _manual_value(
                session,
                "mobile_vehicle",
                "vehicle_used",
                "vehicle",
            ),
            "mileage_start": _manual_value(session, "mileage_start"),
            "mileage_end": _manual_value(session, "mileage_end"),
        }
    ]


def _shift_for_record(record, shifts: list[dict[str, Any]]) -> dict[str, Any]:
    if len(shifts) < 2:
        return shifts[0]

    date_time = pd.to_datetime(record.get("date_time"), errors="coerce")
    if pd.isna(date_time):
        return shifts[0]

    return shifts[0] if date_time.hour < 8 else shifts[1]


def _manual_int(session, *keys: str, default: int = 0) -> int:
    value = _manual_value(session, *keys, default=default)
    number = pd.to_numeric(value, errors="coerce")
    if pd.isna(number):
        return default
    return int(number)


def _transgressions_count(session) -> int:
    value = _manual_value(session, "transgressions_count", "transgressions", default=0)
    if isinstance(value, dict):
        return len(value.get("daily_transgressions") or [])
    if isinstance(value, list):
        return len(value)
    return _manual_int(session, "transgressions_count", default=0)


def _report_date(session, records: pd.DataFrame) -> datetime | str:
    if not records.empty and "date_time" in records:
        dates = records["date_time"].dt.date.dropna().unique().tolist()
        if len(dates) == 1:
            return datetime.combine(dates[0], datetime.min.time())

    try:
        return datetime.strptime(session.report_date, "%Y-%m-%d")
    except ValueError:
        return session.report_date


def _setup_summary_sheet(ws: Worksheet) -> None:
    ws.title = "Weigh & Hourly Summary"
    for column, width in SUMMARY_COLUMN_WIDTHS.items():
        ws.column_dimensions[column].width = width

    ws.row_dimensions[2].height = 15.75
    ws.row_dimensions[3].height = 21.75
    ws.row_dimensions[4].height = 24.75
    for row in range(5, 30):
        ws.row_dimensions[row].height = 15.75 if row in {5, 6, 7, 29} else 15.0

    ws.row_dimensions[30].height = 15.75
    ws.row_dimensions[31].height = 15.75
    ws.row_dimensions[32].height = 55.5
    ws.row_dimensions[33].height = 18.75
    ws.row_dimensions[34].height = 18.75
    ws.row_dimensions[35].height = 19.5
    ws.row_dimensions[36].height = 15.0
    ws.row_dimensions[37].height = 23.25

    _merge_and_set(ws, "C3:C4", "Date", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(ws, "D3:D4", "Time", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(ws, "E3:E4", "Trucks Weighed", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(ws, "F3:F4", "Warned Trucks", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(ws, "G3:G4", "Charged Trucks", bold=True, border=MEDIUM_BORDER)
    _set_cell(ws, "H3", "Excess GVW ", bold=True, border=MEDIUM_BORDER, vertical="top")
    _set_cell(ws, "H4", "Weight(KGS)", bold=True, border=MEDIUM_BORDER, vertical="top")
    _merge_and_set(ws, "J3:L3", "hourly summary", bold=True, border=MEDIUM_BORDER)
    _set_cell(ws, "J4", "Date", bold=True, border=MEDIUM_BORDER)
    _set_cell(ws, "K4", "WEIGHED", bold=True, border=MEDIUM_BORDER)
    _set_cell(ws, "L4", "CHARGED", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(ws, "N3:X30", None, border=MEDIUM_BORDER, wrap_text=False)

    summary_headers = {
        "D32": "Total Weighed (X)",
        "E32": "Warned & Charged (Y)",
        "F32": "Charged & Prohibited (Z)",
        "G32": "Warned (A)",
        "H32": "Cases Cleared in Court (B)",
        "I32": "  Transgressions (L)",
        "J32": "Exempted permit (E)",
        "K32": "Manually Weighed (M)",
    }
    for coordinate, label in summary_headers.items():
        _set_cell(ws, coordinate, label, bold=True, border=MEDIUM_BORDER)

    _merge_and_set(ws, "O32:V32", "Notes", size=28, bold=True, border=MEDIUM_BORDER, wrap_text=False)
    _merge_and_set(
        ws,
        "O33:V33",
        _note_key_value("Red", "formulae, do not edit"),
        size=11,
        bold=False,
        fill=RED_FILL,
        border=MEDIUM_BORDER,
        horizontal="left",
    )
    _merge_and_set(
        ws,
        "O34:V34",
        _note_key_value("No fill", "manual entries fill in from data collection forms"),
        size=11,
        bold=False,
        border=MEDIUM_BORDER,
        horizontal="left",
    )
    _merge_and_set(
        ws,
        "O35:V35",
        "Data in the table is for illustration purposes ONLY",
        size=14,
        bold=True,
        border=MEDIUM_BORDER,
        horizontal="left",
    )


def _setup_detail_sheet(ws: Worksheet, title: str) -> None:
    ws.title = "Mobile Daily Report"
    for column, width in DETAIL_COLUMN_WIDTHS.items():
        ws.column_dimensions[column].width = width

    ws.row_dimensions[1].height = 15.75
    ws.row_dimensions[2].height = 25.5
    ws.row_dimensions[3].height = 14.25
    ws.row_dimensions[4].height = 18.75
    ws.row_dimensions[5].height = 31.5
    ws.row_dimensions[6].height = 18.0
    ws.row_dimensions[7].height = 15.75
    ws.row_dimensions[8].height = 15.75
    ws.row_dimensions[10].height = 30.0

    _merge_and_set(ws, "B2:N2", title, size=14, bold=True, border=THICK_BORDER, fill=LIGHT_GREY_FILL)
    _merge_and_set(ws, "K4:M4", "MILEAGE", bold=True, border=MEDIUM_BORDER, fill=LIGHT_GREY_FILL)

    for coordinate, value in {
        "B5": "DATE",
        "E5": "DANKA STAFF",
        "G5": "POLICE OFFICERS",
        "I5": "TRUCKS WEIGHED",
        "J5": "TRUCKS CHARGED",
        "K5": "MILEAGE START",
        "L5": "MILEAGE END",
        "M5": "KMS",
        "N5": "MOBILE VEHICLE",
    }.items():
        fill = None if coordinate.startswith("B") else LIGHT_GREY_FILL
        _set_cell(ws, coordinate, value, bold=True, border=MEDIUM_BORDER, fill=fill, vertical="bottom")

    _merge_and_set(ws, "C5:D5", "", border=MEDIUM_BORDER, fill=LIGHT_GREY_FILL, vertical="bottom")
    _merge_and_set(ws, "E5:F5", "DANKA STAFF", bold=True, border=MEDIUM_BORDER, fill=LIGHT_GREY_FILL, vertical="bottom")
    _merge_and_set(ws, "G5:H5", "POLICE OFFICERS", bold=True, border=MEDIUM_BORDER, fill=LIGHT_GREY_FILL, vertical="bottom")
    _merge_and_set(ws, "C6:D6", "", border=THIN_BORDER)
    _merge_and_set(ws, "C7:D7", "", border=THIN_BORDER)
    _set_cell(ws, "B7", "", border=THIN_BORDER)

    _set_cell(ws, "G8", "TOTAL", bold=True, border=None)
    _set_cell(ws, "J8", "=J7+J6", bold=True)
    _set_cell(ws, "M8", "=M6", bold=True)

    _merge_and_set(ws, "B9:E9", "", border=MEDIUM_BORDER_THIN_BOTTOM)
    _merge_and_set(ws, "F9:G9", "EXCESS   WEIGHT", bold=True, border=MEDIUM_BORDER_THIN_BOTTOM, fill=LIGHT_GREY_FILL)
    _merge_and_set(ws, "H9:N9", "", border=MEDIUM_BORDER_THIN_BOTTOM)

    headers = {
        "B10": "DATE  WEIGHED",
        "C10": "VEHICLE REG",
        "D10": "TRANSPORTER",
        "E10": "CONFIG.",
        "F10": "GVW",
        "G10": "AXLE",
        "H10": "ORIGIN",
        "I10": "DESTIN.",
        "J10": "CARGO",
        "K10": "COMPUTER OPERATOR \n(DANKA STAFF)",
        "L10": "OFFICERS",
        "M10": "REMARKS",
        "N10": "ROUTE",
    }
    for coordinate, value in headers.items():
        fill = None if coordinate.startswith("B") else LIGHT_GREY_FILL
        _set_cell(
            ws,
            coordinate,
            value,
            bold=True,
            border=THIN_BORDER,
            fill=fill,
            horizontal="center",
            vertical="bottom",
        )


def _write_summary_rows(
    ws: Worksheet,
    records: pd.DataFrame,
    report_date: datetime | str,
    summary: dict[str, Any],
    session=None,
) -> None:
    for offset, hour in enumerate(HOURS):
        row = 5 + offset
        hour_records = records.loc[records["hour_band"].eq(hour)]
        warned = hour_records["remarks"].str.strip().str.upper().eq("WARNED")
        charged = hour_records["is_gvw_axle_charge"] | hour_records["is_dimension_charge"]
        charged_mask = hour_records["remarks"].str.strip().str.upper().eq("CHARGED")
        excess_gvw = hour_records.loc[charged_mask, "gvw_difference_kg"].clip(lower=0).sum()

        _set_cell(
            ws,
            f"C{row}",
            report_date if offset == 0 else None,
            number_format=DATE_FORMAT,
        )
        _set_cell(ws, f"D{row}", hour)
        _set_cell(ws, f"E{row}", int(hour_records["is_weighed"].sum()))
        _set_cell(ws, f"F{row}", int(warned.sum()))
        _set_cell(ws, f"G{row}", int(charged.sum()))
        _set_cell(ws, f"H{row}", int(excess_gvw), number_format="#,##0")
        _set_cell(ws, f"J{row}", hour)
        _set_cell(ws, f"K{row}", f"=E{row}", fill=RED_FILL)
        _set_cell(ws, f"L{row}", f"=G{row}", fill=RED_FILL)

    for coordinate, formula in {
        "D29": "Total",
        "E29": "=SUM(E5:E28)",
        "F29": "=SUM(F5:F28)",
        "G29": "=SUM(G5:G28)",
        "H29": "=SUM(H5:H28)",
        "J29": "Total",
        "K29": "=SUM(K5:K28)",
        "L29": "=SUM(L5:L28)",
    }.items():
        _set_cell(ws, coordinate, formula, bold=True, fill=RED_FILL, border=MEDIUM_BORDER)

    _set_cell(ws, "D33", "=E29", fill=RED_FILL, border=MEDIUM_BORDER)
    _set_cell(ws, "E33", "=F29+G29", fill=RED_FILL, border=MEDIUM_BORDER)
    _set_cell(ws, "F33", "=G29", fill=RED_FILL, border=MEDIUM_BORDER)
    _set_cell(ws, "G33", "=F29", fill=RED_FILL, border=MEDIUM_BORDER)
    _set_cell(ws, "H33", _manual_int(session, "cases_cleared_in_court", default=0), border=MEDIUM_BORDER)
    _set_cell(ws, "I33", _manual_int(session, "transgressions_count", "transgressions", default=0), border=MEDIUM_BORDER)
    _set_cell(ws, "J33", _manual_int(session, "exempted_permit", default=0), border=MEDIUM_BORDER)
    _set_cell(ws, "K33", _manual_int(session, "manually_weighed", default=0), border=MEDIUM_BORDER)
    _merge_and_set(ws, "D34:K34", "weighed summary", bold=True, border=MEDIUM_BORDER)


def _add_hourly_summary_chart(ws: Worksheet) -> None:
    chart = LineChart()
    chart.title = "DAILY HOURLY DATA"
    chart.style = 2
    chart.roundedCorners = False
    chart.varyColors = False
    chart.height = 7.5
    chart.width = 15
    legend_any: Any = chart.legend
    legend_any.position = "b"
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showLegendKey = False
    chart.dataLabels.showVal = False
    chart.dataLabels.showCatName = False
    chart.dataLabels.showSerName = False
    chart.dataLabels.showPercent = False
    chart.dataLabels.showBubbleSize = False

    data = Reference(ws, min_col=11, max_col=12, min_row=4, max_row=28)
    categories = Reference(ws, min_col=10, min_row=5, max_row=28)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    line_colors = (DARK_BLUE_LINE, MAROON_LINE)
    series_list: Any = chart.ser
    for series, color in zip(series_list, line_colors):
        series.smooth = False
        series.marker.symbol = "none"
        series.graphicalProperties.line.width = 28575
        series.graphicalProperties.line.cap = "rnd"
        series.graphicalProperties.line.round = True
        series.graphicalProperties.line.solidFill = color

    x_axis: Any = chart.x_axis
    y_axis: Any = chart.y_axis
    x_axis.axId = 1549586240
    y_axis.axId = 1549567520
    x_axis.crossAx = 1549567520
    y_axis.crossAx = 1549586240
    x_axis.axPos = "b"
    y_axis.axPos = "l"
    x_axis.crosses = "autoZero"
    y_axis.crosses = "autoZero"
    y_axis.crossBetween = "between"
    x_axis.tickLblPos = "nextTo"
    y_axis.tickLblPos = "nextTo"
    x_axis.number_format = "General"
    y_axis.number_format = "General"
    x_axis.textProperties = _chart_axis_text_properties()
    y_axis.textProperties = _chart_axis_text_properties()
    y_axis.spPr = GraphicalProperties(ln=LineProperties(noFill=True))
    x_axis.majorTickMark = "none"
    x_axis.minorTickMark = "none"
    y_axis.majorTickMark = "none"
    y_axis.minorTickMark = "none"

    chart_any: Any = chart
    chart_any.anchor = TwoCellAnchor(
        _from=AnchorMarker(col=13, colOff=728385, row=3, rowOff=51548),
        to=AnchorMarker(col=23, colOff=67238, row=27, rowOff=156883),
    )
    ws.add_chart(chart)


def _write_manual_header(
    ws: Worksheet,
    session,
    report_date: datetime | str,
    summary: dict[str, Any],
) -> None:
    shifts = _manual_shifts(session)

    _set_detail_center_cell(ws, "B6", report_date, number_format=DATE_FORMAT)
    _set_detail_center_cell(ws, "I6", int(summary["total_trucks_weighed"]), border=MEDIUM_BORDER, size=10)
    _set_detail_center_cell(ws, "J6", int(summary["charged_gvw_axle_trucks"]), border=MEDIUM_BORDER, size=10)
    charged_dim = int(summary["charged_dimensions_trucks"])
    _set_detail_center_cell(ws, "J7", charged_dim)

    for index, shift in enumerate(shifts[:2], start=6):
        mileage_start = _plain_int(shift.get("mileage_start"))
        mileage_end = _plain_int(shift.get("mileage_end"))
        _set_detail_text_cell(ws, f"E{index}", _upper(shift.get("danka_staff")))
        _set_detail_text_cell(ws, f"G{index}", _upper(shift.get("police_officers")), vertical=None)
        _set_detail_center_cell(ws, f"K{index}", mileage_start, border=MEDIUM_BORDER)
        _set_detail_center_cell(ws, f"L{index}", mileage_end)
        _set_detail_center_cell(
            ws,
            f"M{index}",
            f"=L{index}-K{index}"
            if mileage_start is not None and mileage_end is not None
            else None,
            border=MEDIUM_BORDER,
            size=10,
        )
        _set_detail_center_cell(
            ws,
            f"N{index}",
            _upper(shift.get("mobile_vehicle")),
            border=MEDIUM_BORDER,
            size=10,
        )

    if len(shifts) < 2:
        _set_detail_text_cell(ws, "E7", "")
        _set_detail_text_cell(ws, "G7", "", vertical=None)


def _write_detail_rows(
    ws: Worksheet,
    records: pd.DataFrame,
    session,
    report_date: datetime | str,
) -> None:
    route = _upper(_manual_value(session, "route"))
    shifts = _manual_shifts(session)

    for offset, (_, record) in enumerate(records.iterrows()):
        row = 11 + offset
        gvw_excess = record["gvw_difference_kg"] if record["gvw_difference_kg"] > 0 else "-"
        axle_excess = record["excess_kg"] if record["excess_kg"] > 0 else "-"

        shift = _shift_for_record(record, shifts)
        danka_staff_val = shift.get("danka_staff")
        police_officers_val = shift.get("police_officers")

        values = {
            f"B{row}": report_date,
            f"C{row}": _upper(record["registration"]),
            f"D{row}": _upper(record["transporter"]),
            f"E{row}": _upper(record["axle"]),
            f"F{row}": int(gvw_excess) if gvw_excess != "-" else "-",
            f"G{row}": int(axle_excess) if axle_excess != "-" else "-",
            f"H{row}": _upper(record["origin"]),
            f"I{row}": _upper(record["destination"]),
            f"J{row}": _upper(record["cargo"]),
            f"K{row}": _upper(danka_staff_val),
            f"L{row}": _upper(police_officers_val),
            f"M{row}": _upper(record["remarks"]),
            f"N{row}": route,
        }

        for coordinate, value in values.items():
            number_format = DATE_FORMAT if coordinate.startswith("B") else "General"
            column = coordinate[0]
            if column in {"B", "F", "G"}:
                _set_detail_center_cell(
                    ws,
                    coordinate,
                    value,
                    number_format=number_format,
                )
            else:
                _set_detail_text_cell(
                    ws,
                    coordinate,
                    value,
                    vertical="center" if column in {"K", "M"} else None,
                    number_format=number_format,
                )


def build_mobile_excel_report(session) -> io.BytesIO:
    if (
        "mobile_report" not in session.dataframes
        or session.sections.get("mobile_report", {}).get("status") != "ready"
    ):
        raise ValueError("Mobile report data is not ready")

    records = session.dataframes["mobile_report"].copy()
    if records.empty:
        raise ValueError("Mobile report data is empty")

    records["date_time"] = pd.to_datetime(records["date_time"], errors="coerce")
    records = records.loc[records["date_time"].notna()].copy()
    if records.empty:
        raise ValueError("Mobile report data has no valid Date Time values")

    summary = summarize_mobile_report(records)
    summary["cases_cleared_in_court"] = _manual_int(
        session,
        "cases_cleared_in_court",
        "cases_cleared_court",
        default=0,
    )
    summary["transgressions_count"] = _transgressions_count(session)
    summary["exempted_permit"] = _manual_int(session, "exempted_permit", default=0)
    summary["manually_weighed"] = _manual_int(session, "manually_weighed", default=0)

    report_date = _report_date(session, records)
    title = f"{session.station} DAILY {session.bound} REPORT".strip().upper()

    wb = Workbook()
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except AttributeError:
        pass

    summary_ws: Any = wb.active
    _setup_summary_sheet(summary_ws)
    _write_summary_rows(summary_ws, records, report_date, summary, session)
    _add_hourly_summary_chart(summary_ws)

    detail_ws: Any = wb.create_sheet("Mobile Daily Report")
    _setup_detail_sheet(detail_ws, title)
    _write_manual_header(detail_ws, session, report_date, summary)
    _write_detail_rows(detail_ws, records, session, report_date)

    hidden_ws: Any = wb.create_sheet("Sheet3")
    hidden_ws.sheet_state = "hidden"

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
