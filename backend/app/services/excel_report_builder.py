import io
from copy import copy
from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.drawing.text import (
    CharacterProperties,
    Paragraph,
    ParagraphProperties,
    RichTextProperties,
)
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from app.services.daily_summary_processor import (
    DailySummaryMissingSourceError,
    build_daily_summary_from_session,
)
from app.services.report_session_metrics import get_wideload_count_from_session


FONT_NAME = "Arial"
NO_FILL = PatternFill(fill_type=None)
YELLOW_FILL = PatternFill("solid", fgColor="FFFFFF00")
RED_FILL = PatternFill("solid", fgColor="FFFF0000")
GREEN_FILL = PatternFill("solid", fgColor="FF00B050")
LIGHT_GRAY_FILL = PatternFill("solid", fgColor="FFD9D9D9")

THIN_SIDE = Side(style="thin", color="000000")
MEDIUM_SIDE = Side(style="medium", color="000000")
NO_SIDE = Side(style=None)

THIN_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)
MEDIUM_BORDER = Border(
    left=MEDIUM_SIDE,
    right=MEDIUM_SIDE,
    top=MEDIUM_SIDE,
    bottom=MEDIUM_SIDE,
)

SUMMARY_COLUMN_WIDTHS = {
    "A": 7.28515625,
    "B": 12.140625,
    "C": 12.85546875,
    "D": 13.0,
    "E": 13.42578125,
    "F": 12.0,
    "G": 12.140625,
    "H": 14.140625,
    "I": 12.85546875,
    "J": 13.140625,
    "K": 12.7109375,
    "L": 13.0,
    "M": 13.140625,
    "N": 12.28515625,
    "O": 13.85546875,
    "P": 11.140625,
    "Q": 10.85546875,
    "R": 10.5703125,
    "S": 9.28515625,
    "T": 11.42578125,
    "U": 14.85546875,
    "V": 15.5703125,
    "W": 16.85546875,
    "X": 15.85546875,
    "Y": 9.140625,
}

SUMMARY_ROW_HEIGHTS = {
    1: 15.75,
    2: 45.75,
    3: 34.5,
    4: 30.75,
    5: 32.25,
    30: 19.5,
    32: 15.0,
    33: 45.0,
    34: 15.0,
    35: 19.5,
    36: 19.5,
    37: 15.0,
    38: 67.5,
    39: 44.25,
    40: 19.5,
    41: 15.0,
    42: 15.0,
    44: 15.0,
    45: 36.0,
    46: 15.75,
    47: 15.75,
    48: 15.75,
    49: 15.0,
}

CC_COLUMN_WIDTHS = {
    "A": 9.140625,
    "B": 17.42578125,
    "C": 21.28515625,
    "D": 21.0,
    "E": 9.140625,
    "F": 9.140625,
    "G": 17.42578125,
    "H": 17.42578125,
    "I": 17.42578125,
    "J": 17.42578125,
}

CC_ROW_HEIGHTS = {
    1: 15.0,
    2: 24.0,
    3: 15.0,
    4: 18.0,
    5: 15.0,
    6: 14.25,
    7: 22.5,
    9: 15.0,
    10: 15.0,
    11: 15.0,
    12: 18.75,
    13: 18.75,
    14: 17.25,
    17: 15.75,
    18: 23.25,
}


def _excel_report_date(report_date: str) -> datetime | str:
    try:
        return datetime.strptime(report_date, "%Y-%m-%d")
    except ValueError:
        return report_date


def _station_bound_title(session) -> str:
    return f"{session.station} {session.bound}".strip().upper()


def _font(size: float = 11, bold: bool = False) -> Font:
    return Font(name=FONT_NAME, size=size, bold=bold)


def _alignment(
    horizontal: str | None = "center",
    vertical: str | None = "center",
    wrap_text: bool | None = True,
) -> Alignment:
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)


def _set_cell(
    ws: Worksheet,
    coordinate: str,
    value: Any = None,
    *,
    size: float = 11,
    bold: bool = False,
    fill: PatternFill | None = NO_FILL,
    border: Border | None = THIN_BORDER,
    horizontal: str | None = "center",
    vertical: str | None = "center",
    wrap_text: bool | None = True,
    number_format: str = "General",
) -> None:
    cell = ws[coordinate]
    cell.value = value
    cell.font = _font(size=size, bold=bold)
    cell.alignment = _alignment(horizontal, vertical, wrap_text)
    if fill is not None:
        cell.fill = copy(fill)
    if border is not None:
        cell.border = copy(border)
    cell.number_format = number_format


def _style_range(
    ws: Worksheet,
    cell_range: str,
    *,
    size: float = 11,
    bold: bool = False,
    fill: PatternFill | None = NO_FILL,
    border: Border | None = THIN_BORDER,
    horizontal: str | None = "center",
    vertical: str | None = "center",
    wrap_text: bool | None = True,
    number_format: str = "General",
) -> None:
    for row in ws[cell_range]:
        for cell in row:
            cell.font = _font(size=size, bold=bold)
            cell.alignment = _alignment(horizontal, vertical, wrap_text)
            if fill is not None:
                cell.fill = copy(fill)
            if border is not None:
                cell.border = copy(border)
            cell.number_format = number_format


def _merge_and_set(
    ws: Worksheet,
    cell_range: str,
    value: Any = None,
    **style_kwargs,
) -> None:
    _style_range(ws, cell_range, **style_kwargs)
    ws.merge_cells(cell_range)
    _set_cell(ws, cell_range.split(":", 1)[0], value, **style_kwargs)


def _numeric_int(value: Any) -> int:
    number = pd.to_numeric(value, errors="coerce")
    if pd.isna(number):
        return 0
    return int(number)


def _daily_rows(daily_df: pd.DataFrame) -> pd.DataFrame:
    if "DATE" not in daily_df.columns:
        return daily_df.head(24).copy()
    mask = daily_df["DATE"].astype(str).str.strip().str.lower().ne("totals")
    return daily_df.loc[mask].head(24).copy()


def _daily_totals(daily_df: pd.DataFrame) -> dict[str, int]:
    if "DATE" in daily_df.columns:
        mask = daily_df["DATE"].astype(str).str.strip().str.lower().eq("totals")
        if mask.any():
            totals_row = daily_df.loc[mask].iloc[-1]
        else:
            totals_row = daily_df.drop(columns=["DATE", "TIME"], errors="ignore").sum(
                numeric_only=True
            )
    else:
        totals_row = daily_df.sum(numeric_only=True)

    return {
        column: _numeric_int(totals_row.get(column, 0))
        for column in ["D", "S", "M", "H", "Q", "X", "C", "Y", "P", "A", "Z", "G", "R", "E"]
    }


def _daily_summary_values(session) -> dict[str, int] | None:
    values = session.sections.get("daily_summary", {}).get("values")
    if values:
        return values

    try:
        return build_daily_summary_from_session(session)
    except DailySummaryMissingSourceError:
        return None


def _traffic_values(session) -> dict[str, int]:
    traffic = session.manual_inputs.get("traffic_census") or {}
    return {
        "buses_gte_3500kg": _numeric_int(traffic.get("buses_gte_3500kg", 0)),
        "vehicles_3500_to_7000_excluding_buses": _numeric_int(
            traffic.get("vehicles_3500_to_7000_excluding_buses", 0)
        ),
        "vehicles_gte_7000_excluding_buses": _numeric_int(
            traffic.get("vehicles_gte_7000_excluding_buses", 0)
        ),
        "total_traffic_census": _numeric_int(traffic.get("total_traffic_census", 0)),
    }


def _set_dimensions(ws: Worksheet, widths: dict[str, float], heights: dict[int, float]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for row in range(6, 30):
        ws.row_dimensions[row].height = 15.0 if row != 7 else 15.75
    for row, height in heights.items():
        ws.row_dimensions[row].height = height


def _split_three_rows(total: int) -> list[int]:
    if total <= 0:
        return [0, 0, 0]

    first = round(total * 0.19)
    second = round(total * 0.55)
    third = total - first - second

    if third < 0:
        third = 0
        second = total - first

    return [first, second, third]


def _apply_sheet_setup(ws: Worksheet) -> None:
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def _write_daily_hour_headers(ws: Worksheet) -> None:
    _merge_and_set(ws, "B2:B5", "  Date", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(ws, "C2:C5", "Time", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(ws, "D2:I2", "Trucks Weighed", bold=True, border=MEDIUM_BORDER)
    _set_cell(ws, "D3", "Multideck", bold=True, border=MEDIUM_BORDER)
    _set_cell(ws, "E3", "Weighed SAW", bold=True, border=MEDIUM_BORDER)
    _set_cell(ws, "F3", "Manually ", bold=True, border=MEDIUM_BORDER)
    _set_cell(ws, "G3", "HSWIM Total", bold=True, border=MEDIUM_BORDER)
    _set_cell(ws, "H3", "HSWIM CLEARED", bold=True, border=MEDIUM_BORDER)
    _set_cell(ws, "I3", "Total weighed", bold=True, border=MEDIUM_BORDER)
    for coordinate, value in {
        "D4:D5": "(D)",
        "E4:E5": "(S)",
        "F4:F5": "(M)",
        "G4:G5": "(H)",
        "H4:H5": "Q = H-C",
        "I4:I5": "X= (D+S+M)",
    }.items():
        _merge_and_set(ws, coordinate, value, bold=True, border=MEDIUM_BORDER)

    _set_cell(ws, "J2", "Called in", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(ws, "J3:J5", "(C)", bold=True, border=MEDIUM_BORDER)
    _set_cell(ws, "K2", "Total overloaded", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(ws, "K3:K5", "(Y)= (A+Z+G+R)", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(ws, "L2:L3", "Impounded &  prohibited ", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(ws, "L4:L5", "(P)= (Z+R)", bold=True, border=MEDIUM_BORDER)
    _set_cell(ws, "M2", "Warned Trucks", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(ws, "M3:M5", "(A)", bold=True, border=MEDIUM_BORDER)
    _set_cell(ws, "N2", "Prohibited & Charged", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(ws, "N3:N5", "(Z)", bold=True, border=MEDIUM_BORDER)
    _set_cell(ws, "O2", "Special Release Trucks", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(ws, "O3:O5", "(G)", bold=True, border=MEDIUM_BORDER)
    _set_cell(ws, "P2", "Redistributed", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(ws, "P3:P5", "(R)", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(ws, "Q2:Q3", "Exemption permits NOT weighed", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(ws, "Q4:Q5", "(E)", bold=True, border=MEDIUM_BORDER)

    _merge_and_set(ws, "T2:T5", "TIME", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(ws, "U2:U4", "Weighed Scale Total (N)", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(ws, "V2:V4", "Manually(M)", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(ws, "W2:W4", "HSWIM CLEARED (Q)", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(ws, "X2:X4", "Total weighed (X)", bold=True, border=MEDIUM_BORDER)
    for coordinate, value in {
        "U5": "N=(D+S)",
        "V5": "(M)",
        "W5": "Q= H-C",
        "X5": "X= (D+S+M)",
    }.items():
        _set_cell(ws, coordinate, value, bold=True, border=MEDIUM_BORDER)


def _write_daily_hour_rows(ws: Worksheet, daily_df: pd.DataFrame, session) -> None:
    rows = _daily_rows(daily_df)
    report_date = _excel_report_date(session.report_date)
    _merge_and_set(
        ws,
        "B6:B29",
        report_date,
        border=Border(left=MEDIUM_SIDE, right=MEDIUM_SIDE, top=THIN_SIDE, bottom=MEDIUM_SIDE),
        vertical="top",
        number_format="mm-dd-yy",
    )

    columns = ["D", "S", "M", "H", "C", "A", "Z", "G", "R", "E"]
    excel_columns = {
        "D": "D",
        "S": "E",
        "M": "F",
        "H": "G",
        "C": "J",
        "A": "M",
        "Z": "N",
        "G": "O",
        "R": "P",
        "E": "Q",
    }

    for offset in range(24):
        excel_row = 6 + offset
        if offset < len(rows):
            row = rows.iloc[offset]
            time_value = row.get("TIME", "")
        else:
            row = {}
            time_value = ""

        _set_cell(ws, f"C{excel_row}", time_value, bold=True, border=THIN_BORDER)
        for source_column in columns:
            coordinate = f"{excel_columns[source_column]}{excel_row}"
            fill = NO_FILL if source_column == "E" else YELLOW_FILL
            _set_cell(
                ws,
                coordinate,
                _numeric_int(row.get(source_column, 0)),
                fill=fill,
                border=THIN_BORDER,
                vertical="top",
                number_format="#,##0" if source_column == "E" else "General",
            )

        formulas = {
            f"H{excel_row}": f"=G{excel_row}-J{excel_row}",
            f"I{excel_row}": f"=D{excel_row}+E{excel_row}+F{excel_row}",
            f"K{excel_row}": f"=M{excel_row}+N{excel_row}+O{excel_row}+P{excel_row}",
            f"L{excel_row}": f"=N{excel_row}+P{excel_row}",
            f"S{excel_row}": f"=D{excel_row}+F{excel_row}",
            f"T{excel_row}": f"=C{excel_row}",
            f"U{excel_row}": f"=D{excel_row}+E{excel_row}",
            f"V{excel_row}": f"=F{excel_row}",
            f"W{excel_row}": f"=H{excel_row}",
            f"X{excel_row}": f"=I{excel_row}",
        }
        for coordinate, formula in formulas.items():
            _set_cell(
                ws,
                coordinate,
                formula,
                size=10 if coordinate.startswith("T") else 11,
                fill=RED_FILL if not coordinate.startswith("S") else NO_FILL,
                border=THIN_BORDER if not coordinate.startswith("S") else None,
                number_format="General",
            )

    _merge_and_set(ws, "B30:C30", "Total", size=12, bold=True, border=MEDIUM_BORDER)
    for column in list("DEFGHIJKLMNOPQ"):
        _set_cell(
            ws,
            f"{column}30",
            f"=SUM({column}6:{column}29)",
            bold=True,
            fill=RED_FILL,
            border=MEDIUM_BORDER,
            number_format="#,##0",
        )
    _set_cell(ws, "T30", "Total", bold=True, fill=RED_FILL, border=MEDIUM_BORDER)
    for column in ["U", "V", "W", "X"]:
        _set_cell(
            ws,
            f"{column}30",
            f"=SUM({column}6:{column}29)",
            bold=True,
            fill=RED_FILL,
            border=MEDIUM_BORDER,
            number_format="#,##0",
        )


def _write_traffic_census_summary(ws: Worksheet, session, totals: dict[str, int]) -> None:
    _merge_and_set(ws, "C33:C35", "Buses>= 3500kg  (J)", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(
        ws,
        "D33:D35",
        "Vehicles>= 3500kg but <7000 excluding buses (V)",
        bold=True,
        border=MEDIUM_BORDER,
    )
    _merge_and_set(
        ws,
        "E33:E35",
        "Vehicles>= 7000 excluding buses (W)",
        bold=True,
        border=MEDIUM_BORDER,
    )
    _merge_and_set(ws, "F33:G33", "Total Traffic Census (K)", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(ws, "F34:G35", " K=(J+V+W) ", bold=True, border=MEDIUM_BORDER)
    _set_cell(ws, "H33", "Exemption\npermits", bold=True, border=MEDIUM_BORDER)
    _set_cell(ws, "I33", "Total Weighed", bold=True, border=MEDIUM_BORDER)
    _set_cell(ws, "J33", "Total Traffic", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(ws, "H34:H35", "(E)", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(ws, "I34:I35", "(X)", bold=True, border=MEDIUM_BORDER)
    _merge_and_set(ws, "J34:J35", "(T)= (Q+X+K+E)", bold=True, border=MEDIUM_BORDER)

    no_left_medium = Border(
        left=NO_SIDE,
        right=MEDIUM_SIDE,
        top=MEDIUM_SIDE,
        bottom=MEDIUM_SIDE,
    )
    _set_cell(
        ws,
        "C36",
        "='CC records'!B14",
        fill=RED_FILL,
        border=MEDIUM_BORDER,
        number_format="#,##0",
    )
    _set_cell(
        ws,
        "D36",
        "='CC records'!C14",
        fill=RED_FILL,
        border=MEDIUM_BORDER,
        number_format="#,##0",
    )
    _set_cell(
        ws,
        "E36",
        "='CC records'!D14",
        fill=RED_FILL,
        border=MEDIUM_BORDER,
        number_format="#,##0",
    )
    _merge_and_set(
        ws,
        "F36:G36",
        "=C36+D36+E36",
        fill=RED_FILL,
        border=no_left_medium,
        number_format="#,##0",
    )
    _set_cell(
        ws,
        "H36",
        "=Q30",
        fill=RED_FILL,
        border=no_left_medium,
        number_format="#,##0",
    )
    _set_cell(
        ws,
        "I36",
        "=I30",
        fill=RED_FILL,
        border=no_left_medium,
        number_format="#,##0",
    )
    _set_cell(
        ws,
        "J36",
        "=F40",
        fill=RED_FILL,
        border=no_left_medium,
        number_format="#,##0",
    )


def _write_daily_summary(ws: Worksheet, session) -> None:
    summary = _daily_summary_values(session) or {
        "weighed_by_hswim_q": 0,
        "weighed_scale_total_n": 0,
        "manually_weighed_m": 0,
        "total_weighed_x": 0,
        "total_traffic_t": 0,
        "total_overload_y": 0,
        "warned_a": 0,
        "charged_prohibited_z": 0,
        "special_release_g": 0,
        "vehicles_charged_but_redistributed_r": 0,
        "impounded_prohibited_p": 0,
        "cases_cleared_in_court_b": 0,
        "transgressions_l": 0,
        "exemption_permits_not_weighed_e": 0,
        "exemption_permits_weighed_f": 0,
        "exemption_permits_total": 0,
    }

    headers_1 = {
        "B38": "HSWIM CLEARED (Q)",
        "C38": "Weighed Scale Total          (N)",
        "D38": "Manually Weighed (M)",
        "E38": "Total weighed (X)",
        "F38:G38": "Total Traffic (T)",
        "H38": "Total   Overload       (Y)            A+Z+G+R",
        "I38": " Warned   (A)",
        "J38": "Charged in court (Z)",
        "K38": "Special release (G)",
        "L38": "Vehicles Charged but Redistributed (R)",
        "M38": "Impounded &  prohibited                (P)= Z+R",
        "N38": "Cases cleared in court (B)",
        "O38": "Transgressions",
        "P38:R38": "Exemption permits",
    }
    for coordinate, value in headers_1.items():
        if ":" in coordinate:
            _merge_and_set(ws, coordinate, value, bold=True, border=MEDIUM_BORDER)
        else:
            _set_cell(ws, coordinate, value, bold=True, border=MEDIUM_BORDER)

    headers_2 = {
        "B39": "(Q=H-C)",
        "C39": "N=(D+S)",
        "D39": "(M)",
        "E39": "(X)=(S+M)",
        "F39:G39": "(T)=(Q+X+K+E)",
        "H39": "(Y)",
        "I39": "(A)",
        "J39": "(Z)",
        "K39": "(G)",
        "L39": "(R)",
        "M39": "(P)",
        "N39": "(B)",
        "O39": "(L)",
        "P39": "NOT weighed (E)",
        "Q39": "Weighed (F)",
        "R39": "Total",
    }
    for coordinate, value in headers_2.items():
        if ":" in coordinate:
            _merge_and_set(ws, coordinate, value, bold=True, border=MEDIUM_BORDER)
        else:
            _set_cell(ws, coordinate, value, bold=True, border=MEDIUM_BORDER)

    no_left_medium = Border(
        left=NO_SIDE,
        right=MEDIUM_SIDE,
        top=MEDIUM_SIDE,
        bottom=MEDIUM_SIDE,
    )
    values = {
        "B40": ("=H30", RED_FILL, Border(left=MEDIUM_SIDE, right=MEDIUM_SIDE, top=NO_SIDE, bottom=MEDIUM_SIDE)),
        "C40": ("=D30+E30", RED_FILL, Border(left=MEDIUM_SIDE, right=MEDIUM_SIDE, top=THIN_SIDE, bottom=MEDIUM_SIDE)),
        "D40": ("=F30", RED_FILL, MEDIUM_BORDER),
        "E40": ("=SUM(C40:D40)", RED_FILL, no_left_medium),
        "F40:G40": ("=E40+F36+H36+B40", RED_FILL, MEDIUM_BORDER),
        "H40": ("=K30", RED_FILL, no_left_medium),
        "I40": ("=M30", RED_FILL, MEDIUM_BORDER),
        "J40": ("=N30", RED_FILL, no_left_medium),
        "K40": ("=O30", RED_FILL, no_left_medium),
        "L40": ("=P30", RED_FILL, no_left_medium),
        "M40": ("=L30", RED_FILL, no_left_medium),
        "N40": (summary["cases_cleared_in_court_b"], NO_FILL, no_left_medium),
        "O40": (summary["transgressions_l"], NO_FILL, no_left_medium),
        "P40": ("=Q30", RED_FILL, MEDIUM_BORDER),
        "Q40": (summary["exemption_permits_weighed_f"], NO_FILL, no_left_medium),
        "R40": ("=P40+Q40", RED_FILL, MEDIUM_BORDER),
    }
    for coordinate, (value, fill, border) in values.items():
        if ":" in coordinate:
            _merge_and_set(
                ws,
                coordinate,
                value,
                fill=fill,
                border=border,
                number_format="#,##0",
                wrap_text=None if coordinate == "R40" else True,
            )
        else:
            _set_cell(
                ws,
                coordinate,
                value,
                fill=fill,
                border=border,
                number_format="#,##0",
            )


def _chart_axis_text_properties() -> RichText:
    return RichText(
        bodyPr=RichTextProperties(
            rot=-60000000,
            spcFirstLastPara=True,
            vertOverflow="ellipsis",
            vert="horz",
            wrap="square",
            anchor="ctr",
            anchorCtr=True,
        ),
        p=[
            Paragraph(
                pPr=ParagraphProperties(
                    defRPr=CharacterProperties(sz=900, b=False)
                )
            )
        ],
    )


def _add_daily_hour_chart(ws: Worksheet) -> None:
    chart = LineChart()
    chart.title = "Graph on Trucks Weighed per Hour"
    chart.style = 2
    chart.x_axis.axId = 1763517568
    chart.y_axis.axId = 1763495520
    chart.x_axis.crossAx = 1763495520
    chart.y_axis.crossAx = 1763517568
    chart.x_axis.axPos = "b"
    chart.y_axis.axPos = "l"
    chart.x_axis.tickLblPos = "nextTo"
    chart.y_axis.tickLblPos = "nextTo"
    chart.x_axis.numFmt = "General"
    chart.y_axis.numFmt = "General"
    chart.x_axis.txPr = _chart_axis_text_properties()
    chart.y_axis.txPr = _chart_axis_text_properties()
    chart.y_axis.scaling.min = -10
    chart.y_axis.scaling.max = 300
    chart.y_axis.majorUnit = 50
    chart.x_axis.crosses = "min"
    chart.y_axis.crosses = "autoZero"
    chart.y_axis.crossBetween = "between"
    chart.y_axis.majorGridlines = ChartLines()
    chart.legend.position = "b"
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showLegendKey = False
    chart.dataLabels.showVal = False
    chart.dataLabels.showCatName = False
    chart.dataLabels.showSerName = False
    chart.dataLabels.showPercent = False
    chart.dataLabels.showBubbleSize = False

    data = Reference(ws, min_col=21, max_col=24, min_row=5, max_row=29)
    categories = Reference(ws, min_col=20, min_row=6, max_row=29)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    colors = ["4472C4", "A0522D", "70AD47", "7030A0"]
    for series, color in zip(chart.series, colors, strict=False):
        series.smooth = False
        series.graphicalProperties.line.solidFill = color
        series.graphicalProperties.line.width = 28575

    chart.anchor = TwoCellAnchor(
        _from=AnchorMarker(col=18, colOff=178594, row=30, rowOff=122283),
        to=AnchorMarker(col=24, colOff=107156, row=41, rowOff=88246),
    )
    ws.add_chart(chart)


def _write_reference_markers(ws: Worksheet) -> None:
    start_row = 48
    ws.row_dimensions[start_row].height = 33.0
    ws.row_dimensions[start_row + 1].height = 15.75
    ws.row_dimensions[start_row + 2].height = 15.75
    ws.row_dimensions[start_row + 3].height = 15.75
    ws.row_dimensions[start_row + 4].height = 15.0

    _merge_and_set(
        ws,
        f"S{start_row}:X{start_row}",
        "Notes",
        size=20,
        bold=True,
        border=MEDIUM_BORDER,
        horizontal="center",
        vertical="center",
        wrap_text=None,
    )
    _merge_and_set(
        ws,
        f"S{start_row + 1}:X{start_row + 1}",
        "yellow = data extracted from Kenload",
        size=12,
        fill=YELLOW_FILL,
        border=MEDIUM_BORDER,
        horizontal="left",
        vertical="center",
        wrap_text=None,
    )
    _merge_and_set(
        ws,
        f"S{start_row + 2}:X{start_row + 2}",
        "red = formulae, do not edit",
        size=12,
        fill=RED_FILL,
        border=MEDIUM_BORDER,
        horizontal="left",
        vertical="center",
        wrap_text=None,
    )
    _merge_and_set(
        ws,
        f"S{start_row + 3}:X{start_row + 3}",
        "No fill = manual entries fill in from data collection forms/books",
        size=12,
        fill=LIGHT_GRAY_FILL,
        border=MEDIUM_BORDER,
        horizontal="left",
        vertical="center",
        wrap_text=None,
    )
    _merge_and_set(
        ws,
        f"S{start_row + 4}:X{start_row + 4}",
        "Data in the table is for illustration purposes ONLY",
        size=12,
        border=None,
        horizontal="left",
        vertical=None,
        wrap_text=None,
    )


def _write_summary_sheet(ws: Worksheet, session) -> None:
    daily_df = session.dataframes["daily_hour"]
    totals = _daily_totals(daily_df)

    _set_dimensions(ws, SUMMARY_COLUMN_WIDTHS, SUMMARY_ROW_HEIGHTS)
    _write_daily_hour_headers(ws)
    _write_daily_hour_rows(ws, daily_df, session)
    _write_traffic_census_summary(ws, session, totals)
    _write_daily_summary(ws, session)
    _write_reference_markers(ws)
    _add_daily_hour_chart(ws)
    _apply_sheet_setup(ws)


def _write_cc_records(ws: Worksheet, session) -> None:
    _set_dimensions(ws, CC_COLUMN_WIDTHS, CC_ROW_HEIGHTS)
    _merge_and_set(ws, "B2:D2", "Total Traffic Census Summary", size=16, bold=True, border=MEDIUM_BORDER, vertical="top")
    _merge_and_set(ws, "B4:D4", _station_bound_title(session), size=14, bold=True, border=Border(left=MEDIUM_SIDE, top=MEDIUM_SIDE, right=MEDIUM_SIDE, bottom=NO_SIDE), wrap_text=None)
    _merge_and_set(ws, "B5:B7", "Buses>= 3500kg  (J)", bold=True, border=THIN_BORDER, vertical="top")
    _merge_and_set(
        ws,
        "C5:C7",
        "Vehicles>= 3500kg but <7000 excluding buses (V)",
        bold=True,
        border=THIN_BORDER,
        vertical="top",
    )
    _merge_and_set(
        ws,
        "D5:D7",
        "Vehicles>= 7000 excluding buses (W)",
        bold=True,
        border=THIN_BORDER,
        vertical="top",
    )

    traffic = _traffic_values(session)
    detailed_records = session.manual_inputs.get("cc_records")
    if isinstance(detailed_records, list) and detailed_records:
        for index, record in enumerate(detailed_records[:6], start=8):
            _set_cell(ws, f"B{index}", _numeric_int(record.get("buses_gte_3500kg", 0)), border=THIN_BORDER)
            _set_cell(
                ws,
                f"C{index}",
                _numeric_int(record.get("vehicles_3500_to_7000_excluding_buses", 0)),
                border=THIN_BORDER,
            )
            _set_cell(
                ws,
                f"D{index}",
                _numeric_int(record.get("vehicles_gte_7000_excluding_buses", 0)),
                border=THIN_BORDER,
            )
    else:
        fallback_rows = zip(
            _split_three_rows(traffic["buses_gte_3500kg"]),
            _split_three_rows(traffic["vehicles_3500_to_7000_excluding_buses"]),
            _split_three_rows(traffic["vehicles_gte_7000_excluding_buses"]),
            strict=True,
        )
        for row_index, values in enumerate(fallback_rows, start=8):
            for column, value in zip(["B", "C", "D"], values, strict=True):
                _set_cell(
                    ws,
                    f"{column}{row_index}",
                    value,
                    border=THIN_BORDER,
                    horizontal=None,
                    vertical=None,
                    wrap_text=None,
                    number_format="General",
                )

    _merge_and_set(ws, "B12:D12", "RELEASED TRUCKS", size=14, fill=YELLOW_FILL, border=Border(left=MEDIUM_SIDE, right=MEDIUM_SIDE, top=MEDIUM_SIDE, bottom=NO_SIDE))
    released_border = Border(left=THIN_SIDE, right=THIN_SIDE, top=NO_SIDE, bottom=THIN_SIDE)
    _set_cell(ws, "B13", 0, border=released_border, vertical=None, wrap_text=None)
    _set_cell(ws, "C13", 0, border=released_border, vertical=None, wrap_text=None)
    _set_cell(ws, "D13", None, border=THIN_BORDER, vertical=None, wrap_text=None)
    _merge_and_set(ws, "G12:J12", "Total Traffic Census (K)", size=14, bold=True, border=MEDIUM_BORDER, vertical="top")
    _merge_and_set(
        ws,
        "G13:J13",
        "=B14+C14+D14",
        fill=GREEN_FILL,
        border=MEDIUM_BORDER,
        number_format="General",
    )

    _set_cell(ws, "B14", "=SUM(B6:B13)", fill=GREEN_FILL, border=MEDIUM_BORDER, number_format="General", vertical=None, wrap_text=None)
    _set_cell(
        ws,
        "C14",
        "=SUM(C6:C13)",
        fill=GREEN_FILL,
        border=MEDIUM_BORDER,
        number_format="General",
        vertical=None,
        wrap_text=None,
    )
    _set_cell(
        ws,
        "D14",
        "=SUM(D6:D13)",
        fill=GREEN_FILL,
        border=MEDIUM_BORDER,
        number_format="General",
        vertical=None,
        wrap_text=None,
    )
    _set_cell(ws, "D17", "                                               ", border=None, horizontal=None, vertical=None, wrap_text=None)
    _apply_sheet_setup(ws)


def _write_hswim_weekly(ws: Worksheet) -> None:
    _apply_sheet_setup(ws)


def build_excel_report(session) -> io.BytesIO:
    if (
        "daily_hour" not in session.dataframes
        or session.sections.get("daily_hour", {}).get("status") != "ready"
    ):
        raise ValueError("Excel report requires ready daily_hour data.")

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    _write_summary_sheet(summary, session)

    cc_records = workbook.create_sheet("CC records")
    _write_cc_records(cc_records, session)

    hswim_weekly = workbook.create_sheet("Hswim Weekly")
    _write_hswim_weekly(hswim_weekly)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
